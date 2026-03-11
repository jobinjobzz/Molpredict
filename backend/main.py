from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, Union, List
import io, httpx, asyncio

# RDKit imports at top level
try:
    from rdkit import Chem
    from rdkit.Chem import Descriptors, Crippen, QED
    from rdkit.Chem import rdMolDescriptors
    from rdkit.Chem import rdMolDescriptors as rdmd
    RDKIT_OK = True
except ImportError:
    RDKIT_OK = False
    print("WARNING: RDKit not available")

app = FastAPI(title="MolPredict API", version="2.1.0")

# ── CORS ───────────────────────────────────────────────────────────────────────
@app.middleware("http")
async def cors_middleware(request: Request, call_next):
    if request.method == "OPTIONS":
        return JSONResponse(status_code=200, content={}, headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type, Authorization",
        })
    response = await call_next(request)
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    return response

# ── Models ─────────────────────────────────────────────────────────────────────
class MoleculeRequest(BaseModel):
    smiles: str
    name: Optional[str] = None

class BatchRequest(BaseModel):
    molecules: List[MoleculeRequest]

class NameLookupRequest(BaseModel):
    query: str

class SimilarityRequest(BaseModel):
    smiles: str
    threshold: Optional[float] = 0.7
    max_results: Optional[int] = 10

class PropertyResult(BaseModel):
    name: str
    value: Union[float, str]
    unit: str
    status: str
    description: str

class ToxicityResult(BaseModel):
    endpoint: str
    probability: float
    status: str
    description: str

class ADMETResult(BaseModel):
    category: str
    endpoint: str
    value: Union[float, str]
    unit: str
    status: str
    description: str

class MoleculeResponse(BaseModel):
    smiles: str
    name: Optional[str]
    valid: bool
    molecular_formula: Optional[str]
    molecular_weight: Optional[float]
    properties: list[PropertyResult]
    lipinski: dict
    toxicity: list[ToxicityResult]
    admet: list[ADMETResult]
    structure_url: Optional[str]
    error: Optional[str] = None

# ── Toxicity ───────────────────────────────────────────────────────────────────
def predict_toxicity(mol):
    try:
        mw=Descriptors.MolWt(mol); logp=Crippen.MolLogP(mol)
        tpsa=Descriptors.TPSA(mol); hbd=rdmd.CalcNumHBD(mol)
        arom=rdmd.CalcNumAromaticRings(mol); hba=rdmd.CalcNumHBA(mol)
        alerts={"nitro":"[N+](=O)[O-]","aniline":"Nc1ccccc1","aldehyde":"[CH]=O",
                "michael":"C=CC=O","epoxide":"C1OC1","halo_ar":"c[F,Cl,Br,I]"}
        found={}
        for k,s in alerts.items():
            try: found[k]=mol.HasSubstructMatch(Chem.MolFromSmarts(s))
            except: found[k]=False
        def clamp(x): return round(min(max(x,0.0),0.95),3)
        h=sum([logp>3 and 0.20,mw>500 and 0.15,found["nitro"] and 0.30,found["aniline"] and 0.20,arom>=3 and 0.15])
        c=sum([logp>3.7 and 0.25,mw>450 and 0.10,arom>=2 and 0.20,hba>=4 and 0.10,found["halo_ar"] and 0.15])
        m=sum([found["nitro"] and 0.40,found["aniline"] and 0.30,found["aldehyde"] and 0.25,found["epoxide"] and 0.35,found["michael"] and 0.20])
        s=sum([found["michael"] and 0.35,found["aldehyde"] and 0.30,found["epoxide"] and 0.30,(2<logp<4) and 0.10])
        a=sum([logp>4 and 0.30,mw>400 and 0.10,arom>=3 and 0.20,found["halo_ar"] and 0.20])
        b=sum([1<=logp<=3 and 0.40,mw<450 and 0.20,tpsa<90 and 0.20,hbd<=3 and 0.10,arom<=2 and 0.10])
        return [
            ToxicityResult(endpoint="Hepatotoxicity",probability=clamp(h),status="bad" if h>0.5 else("warning" if h>0.3 else"good"),description="Estimated liver toxicity risk."),
            ToxicityResult(endpoint="Cardiotoxicity (hERG)",probability=clamp(c),status="bad" if c>0.5 else("warning" if c>0.3 else"good"),description="Estimated hERG channel inhibition risk."),
            ToxicityResult(endpoint="Mutagenicity (Ames)",probability=clamp(m),status="bad" if m>0.4 else("warning" if m>0.2 else"good"),description="Estimated mutagenicity risk."),
            ToxicityResult(endpoint="Skin Sensitization",probability=clamp(s),status="bad" if s>0.4 else("warning" if s>0.2 else"good"),description="Risk of allergic skin reaction."),
            ToxicityResult(endpoint="Aquatic Toxicity",probability=clamp(a),status="bad" if a>0.5 else("warning" if a>0.3 else"good"),description="Environmental toxicity risk."),
            ToxicityResult(endpoint="BBB Penetration",probability=clamp(b),status="warning" if b>0.5 else"good",description="Likelihood of crossing blood-brain barrier."),
        ]
    except Exception as e:
        print(f"Toxicity error: {e}"); return []

# ── ADMET ──────────────────────────────────────────────────────────────────────
def compute_admet(mol):
    try:
        mw=Descriptors.MolWt(mol); logp=Crippen.MolLogP(mol)
        tpsa=Descriptors.TPSA(mol); hbd=rdmd.CalcNumHBD(mol)
        hba=rdmd.CalcNumHBA(mol); rot=rdmd.CalcNumRotatableBonds(mol)
        arom=rdmd.CalcNumAromaticRings(mol)

        lipinski_ok=sum([mw>500,logp>5,hbd>5,hba>10])<=1
        bioavail=1.0 if lipinski_ok and tpsa<90 else(0.5 if tpsa<140 else 0.1)
        bioavail_pct=round(bioavail*100)
        caco2=round(max(0.1,min(50,35-0.3*tpsa+1.2*logp-0.05*mw)),2)
        caco2_s="good" if caco2>10 else("warning" if caco2>2 else"bad")
        pgp=round(min(0.95,max(0.05,0.1+(mw/800)*0.3+(hba/12)*0.3+(arom/5)*0.2)),2)
        pgp_s="bad" if pgp>0.6 else("warning" if pgp>0.4 else"good")
        esol=round(0.16-0.63*logp-0.0062*mw+0.066*rot-0.74*arom,2)
        esol_s="good" if esol>-2 else("warning" if esol>-4 else"bad")
        bbb=round(min(0.95,max(0.05,0.4*(1<=logp<=3)+0.3*(mw<400)+0.2*(tpsa<90)+0.1*(hbd<=3))),2)
        bbb_s="good" if bbb>0.7 else("warning" if bbb>0.4 else"bad")
        vd=round(max(0.04,min(20,0.2+0.5*logp+0.002*mw)),2)
        vd_s="good" if 0.2<=vd<=2 else("warning" if vd<=5 else"bad")
        ppb=round(min(99,max(10,50+8*logp+0.02*mw)),1)
        ppb_s="good" if ppb<90 else("warning" if ppb<95 else"bad")
        cyp3a4_sub=round(min(0.95,max(0.05,0.2+(mw/700)*0.3+(logp/6)*0.2+(arom/4)*0.2+(hba/10)*0.1)),2)
        cyp3a4_sub_s="warning" if cyp3a4_sub>0.5 else"good"
        cyp3a4_inh=round(min(0.90,max(0.05,0.1+(logp/8)*0.35+(arom/5)*0.25+(mw/700)*0.2)),2)
        cyp3a4_inh_s="bad" if cyp3a4_inh>0.6 else("warning" if cyp3a4_inh>0.4 else"good")
        cyp2d6=round(min(0.90,max(0.05,0.05+(logp/7)*0.3+(arom/5)*0.3+(hba/8)*0.1)),2)
        cyp2d6_s="bad" if cyp2d6>0.6 else("warning" if cyp2d6>0.4 else"good")
        t_half=round(max(0.5,min(72,2+logp*3+mw*0.01)),1)
        t_half_s="good" if 4<=t_half<=24 else("warning" if t_half<=48 else"bad")
        renal_cl=round(max(0.1,min(15,8-1.5*logp+0.01*mw)),2)
        renal_s="good" if renal_cl>2 else"warning"
        react_smarts=["C=CC=O","[CH]=O","C1OC1","[N+](=O)[O-]"]
        react_count=sum(1 for s in react_smarts if Chem.MolFromSmarts(s) and mol.HasSubstructMatch(Chem.MolFromSmarts(s)))
        react_risk=round(min(0.9,react_count*0.25),2)
        react_s="bad" if react_risk>0.5 else("warning" if react_risk>0.0 else"good")
        herg_risk=round(min(0.9,max(0.05,0.1+logp*0.08+arom*0.08)),2)
        herg_s="bad" if herg_risk>0.5 else("warning" if herg_risk>0.3 else"good")

        return [
            ADMETResult(category="Absorption",endpoint="Oral Bioavailability",value=bioavail_pct,unit="%",status="good" if bioavail_pct>=70 else("warning" if bioavail_pct>=30 else"bad"),description="Predicted fraction absorbed orally (Lipinski + TPSA model)."),
            ADMETResult(category="Absorption",endpoint="Caco-2 Permeability",value=caco2,unit="×10⁻⁶ cm/s",status=caco2_s,description="Intestinal epithelial permeability. >10 = high, 2–10 = moderate."),
            ADMETResult(category="Absorption",endpoint="P-gp Substrate",value=pgp,unit="prob",status=pgp_s,description="P-glycoprotein efflux pump substrate probability. High = reduced absorption."),
            ADMETResult(category="Absorption",endpoint="Water Solubility (ESOL)",value=esol,unit="log mol/L",status=esol_s,description="Estimated aqueous solubility. >-2 excellent, -2 to -4 moderate."),
            ADMETResult(category="Distribution",endpoint="BBB Penetration",value=bbb,unit="prob",status=bbb_s,description="Blood-brain barrier crossing probability. Relevant for CNS drugs."),
            ADMETResult(category="Distribution",endpoint="Volume of Distribution",value=vd,unit="L/kg",status=vd_s,description="Apparent VD at steady state. 0.2–2 L/kg typical for many drugs."),
            ADMETResult(category="Distribution",endpoint="Plasma Protein Binding",value=ppb,unit="%",status=ppb_s,description="Fraction bound to plasma proteins. >95% = highly bound."),
            ADMETResult(category="Metabolism",endpoint="CYP3A4 Substrate",value=cyp3a4_sub,unit="prob",status=cyp3a4_sub_s,description="Probability of being metabolised by CYP3A4 (most abundant liver CYP)."),
            ADMETResult(category="Metabolism",endpoint="CYP3A4 Inhibitor",value=cyp3a4_inh,unit="prob",status=cyp3a4_inh_s,description="Risk of inhibiting CYP3A4, causing drug-drug interactions."),
            ADMETResult(category="Metabolism",endpoint="CYP2D6 Inhibitor",value=cyp2d6,unit="prob",status=cyp2d6_s,description="Risk of inhibiting CYP2D6, important for CNS/cardiac drugs."),
            ADMETResult(category="Metabolism",endpoint="Half-life (t½)",value=t_half,unit="h",status=t_half_s,description="Estimated plasma half-life. 4–24 h is ideal for oral drugs."),
            ADMETResult(category="Excretion",endpoint="Renal Clearance",value=renal_cl,unit="mL/min/kg",status=renal_s,description="Estimated renal clearance. Low logP molecules tend to be renally excreted."),
            ADMETResult(category="Toxicity",endpoint="hERG Inhibition",value=herg_risk,unit="prob",status=herg_s,description="hERG potassium channel blockade risk — major cause of cardiac toxicity."),
            ADMETResult(category="Toxicity",endpoint="Reactive Metabolite Risk",value=react_risk,unit="prob",status=react_s,description="Risk of forming reactive metabolites (epoxides, aldehydes, Michael acceptors)."),
        ]
    except Exception as e:
        print(f"ADMET error: {e}"); return []

# ── Core computation ───────────────────────────────────────────────────────────
def compute_properties(smiles: str) -> MoleculeResponse:
    if not RDKIT_OK:
        return MoleculeResponse(smiles=smiles,name=None,valid=False,molecular_formula=None,
            molecular_weight=None,properties=[],lipinski={},toxicity=[],admet=[],structure_url=None,error="RDKit not installed.")
    mol=Chem.MolFromSmiles(smiles)
    if mol is None:
        return MoleculeResponse(smiles=smiles,name=None,valid=False,molecular_formula=None,
            molecular_weight=None,properties=[],lipinski={},toxicity=[],admet=[],structure_url=None,error="Invalid SMILES string.")
    formula=rdmd.CalcMolFormula(mol)
    mw=Descriptors.MolWt(mol); hbd=rdmd.CalcNumHBD(mol); hba=rdmd.CalcNumHBA(mol)
    logp=Crippen.MolLogP(mol); tpsa=Descriptors.TPSA(mol)
    rot=rdmd.CalcNumRotatableBonds(mol); arom=rdmd.CalcNumAromaticRings(mol)
    heavy=mol.GetNumHeavyAtoms(); qed=QED.qed(mol)
    viol=sum([mw>500,logp>5,hbd>5,hba>10])
    properties=[
        PropertyResult(name="Molecular Weight",value=round(mw,2),unit="Da",status="good" if mw<=500 else("warning" if mw<=600 else"bad"),description="Lipinski limit ≤ 500 Da."),
        PropertyResult(name="LogP (Lipophilicity)",value=round(logp,3),unit="",status="good" if -0.5<=logp<=5 else("warning" if logp<=6 else"bad"),description="Lipinski limit ≤ 5."),
        PropertyResult(name="H-Bond Donors",value=hbd,unit="",status="good" if hbd<=5 else"bad",description="Lipinski limit ≤ 5."),
        PropertyResult(name="H-Bond Acceptors",value=hba,unit="",status="good" if hba<=10 else"bad",description="Lipinski limit ≤ 10."),
        PropertyResult(name="TPSA",value=round(tpsa,2),unit="Å²",status="good" if tpsa<=90 else("warning" if tpsa<=140 else"bad"),description="< 90 Å² good oral absorption."),
        PropertyResult(name="Rotatable Bonds",value=rot,unit="",status="good" if rot<=10 else("warning" if rot<=15 else"bad"),description="Measures molecular flexibility."),
        PropertyResult(name="Aromatic Rings",value=arom,unit="",status="good" if arom<=3 else("warning" if arom<=4 else"bad"),description="More rings increase toxicity risk."),
        PropertyResult(name="QED Score",value=round(qed,3),unit="",status="good" if qed>=0.6 else("warning" if qed>=0.4 else"bad"),description="Drug-likeness 0–1, higher is better."),
        PropertyResult(name="Heavy Atom Count",value=heavy,unit="",status="good" if heavy<=30 else("warning" if heavy<=40 else"bad"),description="Non-hydrogen atoms."),
    ]
    lipinski={"pass":viol<=1,"violations":viol,"mw_ok":mw<=500,"logp_ok":logp<=5,"hbd_ok":hbd<=5,"hba_ok":hba<=10}
    from urllib.parse import quote
    structure_url=f"https://cactus.nci.nih.gov/chemical/structure/{quote(smiles)}/image"
    return MoleculeResponse(smiles=smiles,name=None,valid=True,molecular_formula=formula,
        molecular_weight=round(mw,2),properties=properties,lipinski=lipinski,
        toxicity=predict_toxicity(mol),admet=compute_admet(mol),structure_url=structure_url)

# ── Excel export ───────────────────────────────────────────────────────────────
def results_to_excel(results) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws1=wb.active; ws1.title="Summary"
    hf=PatternFill("solid",fgColor="0D1B2A"); af=PatternFill("solid",fgColor="0A1520")
    hfont=Font(color="60A5FA",bold=True,size=10); wfont=Font(color="E2E8F0",size=10)
    gfont=Font(color="4ADE80",bold=True,size=10); yfont=Font(color="FCD34D",bold=True,size=10)
    rfont=Font(color="F87171",bold=True,size=10)
    ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
    thin=Border(left=Side(style="thin",color="1E3A5C"),right=Side(style="thin",color="1E3A5C"),
                top=Side(style="thin",color="1E3A5C"),bottom=Side(style="thin",color="1E3A5C"))
    def sf(s): return gfont if s=="good" else(yfont if s=="warning" else rfont)
    pn=["Molecular Weight","LogP (Lipophilicity)","H-Bond Donors","H-Bond Acceptors",
        "TPSA","Rotatable Bonds","Aromatic Rings","QED Score","Heavy Atom Count"]
    tn=["Hepatotoxicity","Cardiotoxicity (hERG)","Mutagenicity (Ames)","Skin Sensitization","Aquatic Toxicity","BBB Penetration"]
    an=["Oral Bioavailability","Caco-2 Permeability","P-gp Substrate","Water Solubility (ESOL)",
        "BBB Penetration","Volume of Distribution","Plasma Protein Binding",
        "CYP3A4 Substrate","CYP3A4 Inhibitor","CYP2D6 Inhibitor","Half-life (t½)",
        "Renal Clearance","hERG Inhibition","Reactive Metabolite Risk"]
    def make_sheet(ws,headers):
        for ci,h in enumerate(headers,1):
            c=ws.cell(row=1,column=ci,value=h)
            c.fill=hf; c.font=hfont; c.alignment=ctr; c.border=thin
        ws.row_dimensions[1].height=28
    # Sheet 1
    make_sheet(ws1,["#","Name","SMILES","Formula","Lipinski","Violations"]+pn+[t+" (%)" for t in tn])
    for ri,r in enumerate(results,2):
        rf=af if ri%2==0 else PatternFill("solid",fgColor="060D14")
        col=1
        def wc(val,font=None):
            nonlocal col
            c=ws1.cell(row=ri,column=col,value=val)
            c.fill=rf; c.font=font or wfont; c.alignment=ctr; c.border=thin; col+=1
        wc(ri-1); wc(r.name or "—")
        c2=ws1.cell(row=ri,column=col,value=r.smiles)
        c2.fill=rf; c2.font=wfont; c2.border=thin; c2.alignment=Alignment(horizontal="left",vertical="center"); col+=1
        wc(r.molecular_formula or "—")
        lp=r.lipinski.get("pass",False); wc("PASS" if lp else"FAIL",font=gfont if lp else rfont)
        v=r.lipinski.get("violations","—"); wc(v,font=gfont if v==0 else(yfont if v==1 else rfont))
        pm={p.name:p for p in r.properties}
        for p_name in pn:
            p=pm.get(p_name); wc(f"{p.value} {p.unit}".strip() if p else"—",font=sf(p.status) if p else wfont)
        tm={t.endpoint:t for t in r.toxicity}
        for t_name in tn:
            t=tm.get(t_name); wc(round(t.probability*100,1) if t else"—",font=sf(t.status) if t else wfont)
    for i,w in enumerate([4,16,40,12,10,10]+[16]*len(pn)+[22]*len(tn),1):
        ws1.column_dimensions[get_column_letter(i)].width=w
    # Sheet 2: ADMET
    ws2=wb.create_sheet("ADMET Profile"); make_sheet(ws2,["#","Name"]+an)
    for ri,r in enumerate(results,2):
        rf=af if ri%2==0 else PatternFill("solid",fgColor="060D14")
        ws2.cell(row=ri,column=1,value=ri-1).fill=rf
        ws2.cell(row=ri,column=2,value=r.name or "—").fill=rf
        am={a.endpoint:a for a in r.admet}
        for ci,ep in enumerate(an,3):
            a=am.get(ep); c=ws2.cell(row=ri,column=ci,value=f"{a.value} {a.unit}".strip() if a else"—")
            c.fill=rf; c.font=sf(a.status) if a else wfont; c.alignment=ctr; c.border=thin
    for i,w in enumerate([4,16]+[20]*len(an),1): ws2.column_dimensions[get_column_letter(i)].width=w
    # Sheet 3: Properties
    ws3=wb.create_sheet("Properties"); make_sheet(ws3,["#","Name","SMILES"]+pn)
    for ri,r in enumerate(results,2):
        rf=af if ri%2==0 else PatternFill("solid",fgColor="060D14")
        for ci,val in enumerate([ri-1,r.name or"—",r.smiles],1): ws3.cell(row=ri,column=ci,value=val).fill=rf
        pm={p.name:p for p in r.properties}
        for ci,p_name in enumerate(pn,4):
            p=pm.get(p_name); c=ws3.cell(row=ri,column=ci,value=f"{p.value} {p.unit}".strip() if p else"—")
            c.fill=rf; c.font=sf(p.status) if p else wfont; c.alignment=ctr; c.border=thin
    for i,w in enumerate([4,16,40]+[16]*len(pn),1): ws3.column_dimensions[get_column_letter(i)].width=w
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.get("/")
def root(): return {"message":"MolPredict API","version":"2.0.0"}

@app.get("/health")
def health(): return {"status":"ok"}

@app.get("/examples")
def examples():
    return [
        {"name":"Aspirin","smiles":"CC(=O)Oc1ccccc1C(=O)O"},
        {"name":"Ibuprofen","smiles":"CC(C)Cc1ccc(cc1)C(C)C(=O)O"},
        {"name":"Caffeine","smiles":"Cn1cnc2c1c(=O)n(c(=O)n2C)C"},
        {"name":"Paracetamol","smiles":"CC(=O)Nc1ccc(O)cc1"},
        {"name":"Penicillin G","smiles":"CC1(C)SC2C(NC(=O)Cc3ccccc3)C(=O)N2C1C(=O)O"},
        {"name":"Metformin","smiles":"CN(C)C(=N)NC(=N)N"},
        {"name":"Sildenafil","smiles":"CCCC1=NN(C)C(=C1C(=O)c1ccc(cc1)S(=O)(=O)N1CCN(CC1)C)c1nc(cc(=O)[nH]1)c1cccc(c1)OCC"},
    ]

@app.post("/predict",response_model=MoleculeResponse)
def predict(req: MoleculeRequest):
    result=compute_properties(req.smiles.strip())
    if req.name: result.name=req.name
    return result

@app.post("/batch")
def batch_predict(req: BatchRequest):
    results=[]
    for m in req.molecules:
        r=compute_properties(m.smiles.strip())
        if m.name: r.name=m.name
        results.append(r)
    return results

@app.post("/batch/export")
def batch_export(req: BatchRequest):
    results=[]
    for m in req.molecules:
        r=compute_properties(m.smiles.strip())
        if m.name: r.name=m.name
        results.append(r)
    excel_bytes=results_to_excel(results)
    return StreamingResponse(io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=molpredict_results.xlsx"})

# ── Shared helpers ─────────────────────────────────────────────────────────────
async def _pubchem_name_to_smiles(client, query: str):
    """PubChem: name → SMILES"""
    from urllib.parse import quote
    try:
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{quote(query)}/property/IsomericSMILES,IUPACName,MolecularFormula,MolecularWeight/JSON"
        r = await client.get(url)
        if r.status_code == 200:
            props = r.json()["PropertyTable"]["Properties"][0]
            smiles = props.get("IsomericSMILES", "")
            if smiles:
                return {"found": True, "smiles": smiles, "source": "PubChem",
                        "iupac_name": props.get("IUPACName",""), "formula": props.get("MolecularFormula",""),
                        "mw": str(props.get("MolecularWeight","")), "ext_id": str(props.get("CID",""))}
    except Exception as e:
        print(f"PubChem lookup error: {e}")
    return None

async def _chembl_name_to_smiles(client, query: str):
    """ChEMBL: name/synonym → SMILES"""
    from urllib.parse import quote
    try:
        # Search by preferred name
        url = f"https://www.ebi.ac.uk/chembl/api/data/molecule.json?pref_name__iexact={quote(query)}&limit=1"
        r = await client.get(url)
        if r.status_code == 200:
            mols = r.json().get("molecules", [])
            if mols:
                m = mols[0]
                smiles = m.get("molecule_structures", {}).get("canonical_smiles", "")
                if smiles:
                    return {"found": True, "smiles": smiles, "source": "ChEMBL",
                            "iupac_name": m.get("pref_name",""), "formula": m.get("molecule_properties",{}).get("full_molecular_formula",""),
                            "mw": str(m.get("molecule_properties",{}).get("full_mwt","")),
                            "ext_id": m.get("molecule_chembl_id",""),
                            "max_phase": m.get("max_phase", 0)}
        # Fallback: synonym search
        url2 = f"https://www.ebi.ac.uk/chembl/api/data/molecule.json?molecule_synonyms__synonym__iexact={quote(query)}&limit=1"
        r2 = await client.get(url2)
        if r2.status_code == 200:
            mols2 = r2.json().get("molecules", [])
            if mols2:
                m = mols2[0]
                smiles = m.get("molecule_structures", {}).get("canonical_smiles", "")
                if smiles:
                    return {"found": True, "smiles": smiles, "source": "ChEMBL",
                            "iupac_name": m.get("pref_name",""), "formula": m.get("molecule_properties",{}).get("full_molecular_formula",""),
                            "mw": str(m.get("molecule_properties",{}).get("full_mwt","")),
                            "ext_id": m.get("molecule_chembl_id",""),
                            "max_phase": m.get("max_phase", 0)}
    except Exception as e:
        print(f"ChEMBL lookup error: {e}")
    return None

async def _unichem_name_to_smiles(client, query: str):
    """UniChem: name → InChIKey → SMILES via cross-reference"""
    from urllib.parse import quote
    try:
        # UniChem compound search
        url = f"https://www.ebi.ac.uk/unichem/api/v1/compounds?name={quote(query)}"
        r = await client.get(url)
        if r.status_code == 200:
            data = r.json()
            compounds = data.get("compounds", [])
            if compounds:
                inchikey = compounds[0].get("standardInchiKey", "")
                if inchikey:
                    # Use InChIKey to get SMILES from PubChem
                    r2 = await client.get(f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/inchikey/{inchikey}/property/IsomericSMILES/JSON")
                    if r2.status_code == 200:
                        props = r2.json()["PropertyTable"]["Properties"][0]
                        smiles = props.get("IsomericSMILES","")
                        if smiles:
                            return {"found": True, "smiles": smiles, "source": "UniChem",
                                    "iupac_name": query, "formula": "", "mw": "", "ext_id": inchikey}
    except Exception as e:
        print(f"UniChem lookup error: {e}")
    return None

async def _zinc_name_to_smiles(client, query: str):
    """ZINC: name search via their API"""
    from urllib.parse import quote
    try:
        url = f"https://zinc.docking.org/substances.json?name={quote(query)}&count=1"
        r = await client.get(url, headers={"Accept": "application/json"})
        if r.status_code == 200:
            items = r.json()
            if items and len(items) > 0:
                smiles = items[0].get("smiles","")
                if smiles:
                    return {"found": True, "smiles": smiles, "source": "ZINC",
                            "iupac_name": items[0].get("name", query),
                            "formula": "", "mw": str(items[0].get("mwt","")),
                            "ext_id": items[0].get("zinc_id","")}
    except Exception as e:
        print(f"ZINC lookup error: {e}")
    return None


@app.post("/lookup")
async def name_lookup(req: NameLookupRequest):
    query = req.query.strip()

    # ── Tier 1: Hardcoded common drugs (instant) ───────────────────────────────
    COMMON = {
        "aspirin": "CC(=O)Oc1ccccc1C(=O)O",
        "ibuprofen": "CC(C)Cc1ccc(cc1)C(C)C(=O)O",
        "caffeine": "Cn1cnc2c1c(=O)n(c(=O)n2C)C",
        "paracetamol": "CC(=O)Nc1ccc(O)cc1",
        "acetaminophen": "CC(=O)Nc1ccc(O)cc1",
        "metformin": "CN(C)C(=N)NC(=N)N",
        "sildenafil": "CCCC1=NN(C)C(=C1C(=O)c1ccc(cc1)S(=O)(=O)N1CCN(CC1)C)c1nc(cc(=O)[nH]1)c1cccc(c1)OCC",
        "atorvastatin": "CC(C)c1c(C(=O)Nc2ccccc2)c(-c2ccccc2)c(-c2ccc(F)cc2)n1CCC(O)CC(O)CC(=O)O",
        "omeprazole": "COc1ccc2[nH]c(S(=O)Cc3ncc(C)c(OC)c3C)nc2c1",
        "amoxicillin": "CC1(C)SC2C(NC(=O)C(N)c3ccccc3)C(=O)N2C1C(=O)O",
        "penicillin": "CC1(C)SC2C(NC(=O)Cc3ccccc3)C(=O)N2C1C(=O)O",
        "morphine": "OC1=CC=C2CC3N(C)CCC34C2=C1OC4",
        "glucose": "OC[C@H]1OC(O)[C@H](O)[C@@H](O)[C@@H]1O",
        "dopamine": "NCCc1ccc(O)c(O)c1",
        "serotonin": "NCCc1c[nH]c2ccc(O)cc12",
        "adrenaline": "CNC[C@@H](O)c1ccc(O)c(O)c1",
        "epinephrine": "CNC[C@@H](O)c1ccc(O)c(O)c1",
        "cholesterol": "CC(C)CCCC(C)C1CCC2C3CC=C4CC(O)CCC4(C)C3CCC12C",
        "testosterone": "CC12CCC3C(C1CCC2=O)CCC4=CC(=O)CCC34C",
        "estradiol": "OC1=CC2=C(CC[C@@H]3[C@@H]2CC[C@]4(C)[C@@H]3CC[C@@H]4O)C=C1",
        "warfarin": "CC(=O)CC(c1ccccc1)c1c(O)c2ccccc2oc1=O",
        "paclitaxel": "O=C(O[C@@H]1C[C@]2(O)C(=O)[C@@H](OC(=O)c3ccccc3)[C@@]3(C)[C@H](OC(C)=O)[C@@H](O)[C@H](c4ccccc4)[C@@H]3[C@@H]2CC1)c1ccccc1",
        "cisplatin": "[NH3][Pt](Cl)(Cl)[NH3]",
        "doxorubicin": "COc1cccc2C(=O)c3c(O)c4C[C@@](O)(C[C@H](O[C@H]5C[C@H](N)[C@H](O)[C@H](C)O5)c4c(O)c3C(=O)c12)C(=O)CO",
        "cyclophosphamide": "ClCCN(CCCl)P1(=O)NCCCO1",
        "fluoxetine": "CNCCC(Oc1ccc(cc1)C(F)(F)F)c1ccccc1",
        "diazepam": "ClC1=CC2=C(C=C1)N(C)C(=O)CN=C2c1ccccc1",
        "alprazolam": "Cc1nnc2CN=C(c3ccccc3)c3cc(Cl)ccc3-n12",
        "lisinopril": "NCCCC[C@H](N[C@@H](CCc1ccccc1)C(=O)O)C(=O)N1CCC[C@H]1C(=O)O",
        "amlodipine": "CCOC(=O)C1=C(COCCN)NC(C)=C(C1c1ccccc1Cl)C(=O)OC",
        "simvastatin": "CCC(C)(C)C(=O)O[C@H]1C[C@@H](C)C=C2C=C[C@H](C)[C@H](CC[C@@H]3C[C@@H](O)CC(=O)O3)[C@H]12",
        "losartan": "CCCCc1nc(Cl)c(CO)n1Cc1ccc(-c2ccccc2-c2nnn[nH]2)cc1",
        "metoprolol": "COCCc1ccc(OCC(O)CNC(C)C)cc1",
        "gabapentin": "NCC1(CC(=O)O)CCCCC1",
        "pregabalin": "CC(CN)CC(=O)O",
        "tamoxifen": "CCC(=C(CC)c1ccc(OCCN(C)C)cc1)c1ccccc1",
        "naloxone": "O=C1CC[C@]2(O)CC[N@@+]3(CC[C@H](O2)[C@H]13)CC=C",
        "naltrexone": "O=C1CC[C@]2(O)CC[N@@+]3(CC[C@H](O2)[C@H]13)CC1CC1",
        "fentanyl": "CCC(=O)N(c1ccccc1)C1CCN(CCc2ccccc2)CC1",
        "codeine": "COc1ccc2CC3N(C)CCC34c2c1OC4",
        "lidocaine": "CCN(CC)CC(=O)Nc1c(C)cccc1C",
        "propofol": "CC(C)c1cccc(C(C)C)c1O",
        "ketamine": "O=C1CCCCC1=NC1CCCCC1",
        "haloperidol": "OC1(CCc2ccc(Cl)cc2)CCN(CCCC(=O)c2ccc(F)cc2)CC1",
        "clozapine": "CN1CCN(CC1)c1nc2cc(Cl)ccc2nc2ccccc12",
        "risperidone": "Cc1nc2ccc(F)cc2c(=O)n1CCCC1CCN2CC(=O)Nc2c1",
        "quetiapine": "O=C1CCCCN1CCCN1CCN(c2nc3ccccc3sc2)CC1",
        "fluconazole": "OC(Cn1ccnc1)(Cn1ccnc1)c1ccc(F)cc1F",
        "itraconazole": "CCC(C)n1ncn(c1=O)c1ccc(cc1)N1CCN(CC1)c1ccc(OCC2COc3ccccc3O2)cc1",
        "amphotericin": "OC1C=CC=CC=CC=CC=CC=CC(CC(O)CC(O)CC(OC2OC(C(O)C(O)C2O)C(N)=O)CC(O)CC1=O)C(=O)O",
        "metronidazole": "Cc1ncc([N+](=O)[O-])n1CCO",
        "chloroquine": "CCN(CC)CCCC(C)Nc1ccnc2cc(Cl)ccc12",
        "hydroxychloroquine": "CCN(CCO)CCCC(C)Nc1ccnc2cc(Cl)ccc12",
        "remdesivir": "CCC(CC)COC(=O)[C@@H](N[P@@](=O)(OC[C@H]1O[C@@](C#N)(c2ccc3[nH]cnc3n2)[C@H](O)[C@@H]1F)Oc1ccccc1)C",
        "oseltamivir": "CCOC(=O)[C@@H]1CC(=C[C@H](N)[C@@H]1OC(=O)CC)N",
        "azithromycin": "CC[C@@H]1OC(=O)[C@H](C)[C@@H](O[C@H]2C[C@@](C)(OC)[C@@H](O)[C@H](C)O2)[C@H](C)[C@@H](O[C@@H]2O[C@H](CC)[C@@H](O[C@H]3C[C@@H](N(C)C)[C@@H](O)[C@H](C)O3)[C@H](C)O2)[C@@H](C)C[C@@](O)(CC)[C@@H](C)C(=O)[C@H](C)[C@@H]1N(C)C",
    }

    q_lower = query.lower()
    if q_lower in COMMON:
        print(f"Lookup hit (local): {query}")
        return {"found": True, "smiles": COMMON[q_lower], "query": query,
                "source": "Local DB", "sources_tried": ["Local DB"]}

    # ── Tier 2: Query all databases in parallel ────────────────────────────────
    sources_tried = []
    async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
        tasks = {
            "pubchem": _pubchem_name_to_smiles(client, query),
            "chembl":  _chembl_name_to_smiles(client, query),
            "unichem": _unichem_name_to_smiles(client, query),
            "zinc":    _zinc_name_to_smiles(client, query),
        }
        results_raw = await asyncio.gather(*tasks.values(), return_exceptions=True)
        db_results = dict(zip(tasks.keys(), results_raw))

        # Return first successful hit, in priority order
        for db in ["pubchem", "chembl", "unichem", "zinc"]:
            sources_tried.append(db.capitalize())
            res = db_results.get(db)
            if res and isinstance(res, dict) and res.get("found") and res.get("smiles"):
                print(f"Lookup hit ({res['source']}): {query}")
                res["query"] = query
                res["sources_tried"] = sources_tried
                return res

    print(f"Lookup miss for '{query}' across all databases")
    return {"found": False, "query": query, "smiles": "",
            "sources_tried": sources_tried,
            "message": f"'{query}' not found in PubChem, ChEMBL, UniChem, or ZINC. Try the full IUPAC or generic name."}


@app.post("/similarity")
async def similarity_search(req: SimilarityRequest):
    from urllib.parse import quote
    smiles = req.smiles.strip()
    threshold = int((req.threshold or 0.7) * 100)
    max_res = min(req.max_results or 10, 20)

    async def pubchem_similarity(client):
        """PubChem 2D fingerprint similarity — 100M+ compounds"""
        try:
            r = await client.get(
                f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/fastsimilarity_2d/smiles/"
                f"{quote(smiles)}/cids/JSON?Threshold={threshold}&MaxRecords={max_res}"
            )
            if r.status_code != 200: return []
            cids = r.json().get("IdentifierList", {}).get("CID", [])
            if not cids: return []
            cid_str = ",".join(str(c) for c in cids[:max_res])
            rp = await client.get(
                f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid_str}"
                f"/property/IsomericSMILES,IUPACName,MolecularFormula,MolecularWeight,XLogP,TPSA/JSON"
            )
            if rp.status_code != 200: return []
            props = rp.json().get("PropertyTable", {}).get("Properties", [])
            # Get synonyms
            rs = await client.get(f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid_str}/synonyms/JSON")
            syn_map = {}
            if rs.status_code == 200:
                for item in rs.json().get("InformationList", {}).get("Information", []):
                    cid = item.get("CID"); syns = item.get("Synonym", [])
                    syn_map[cid] = next((s for s in syns if len(s) <= 30 and s[0].isupper()), syns[0] if syns else f"CID {cid}")
            return [{
                "name": syn_map.get(p.get("CID"), f"CID {p.get('CID')}"),
                "smiles": p.get("IsomericSMILES", ""),
                "formula": p.get("MolecularFormula", ""),
                "mw": str(p.get("MolecularWeight", "")),
                "xlogp": str(p.get("XLogP", "")),
                "tpsa": str(p.get("TPSA", "")),
                "source": "PubChem",
                "source_color": "#3b82f6",
                "ext_id": str(p.get("CID", "")),
                "pubchem_url": f"https://pubchem.ncbi.nlm.nih.gov/compound/{p.get('CID')}",
                "structure_url": f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{p.get('CID')}/PNG",
            } for p in props if p.get("IsomericSMILES")]
        except Exception as e:
            print(f"PubChem similarity error: {e}"); return []

    async def chembl_similarity(client):
        """ChEMBL structural similarity search — approved drugs + bioactivity"""
        try:
            r = await client.get(
                f"https://www.ebi.ac.uk/chembl/api/data/similarity/{quote(smiles)}/{threshold}.json?limit={max_res}"
            )
            if r.status_code != 200: return []
            mols = r.json().get("molecules", [])
            results = []
            for m in mols:
                s = m.get("molecule_structures", {}).get("canonical_smiles", "")
                if not s: continue
                props = m.get("molecule_properties") or {}
                chembl_id = m.get("molecule_chembl_id", "")
                results.append({
                    "name": m.get("pref_name") or chembl_id,
                    "smiles": s,
                    "formula": props.get("full_molecular_formula", ""),
                    "mw": str(props.get("full_mwt", "")),
                    "xlogp": str(props.get("alogp", "")),
                    "tpsa": str(props.get("psa", "")),
                    "source": "ChEMBL",
                    "source_color": "#f59e0b",
                    "ext_id": chembl_id,
                    "max_phase": m.get("max_phase", 0),
                    "pubchem_url": f"https://www.ebi.ac.uk/chembl/compound_report_card/{chembl_id}/",
                    "structure_url": f"https://www.ebi.ac.uk/chembl/api/data/image/{chembl_id}.svg",
                })
            return results
        except Exception as e:
            print(f"ChEMBL similarity error: {e}"); return []

    async def zinc_similarity(client):
        """ZINC purchasable compound search by substructure/similarity"""
        try:
            r = await client.get(
                f"https://zinc.docking.org/substances.json?smiles={quote(smiles)}&count={max_res}",
                headers={"Accept": "application/json"}
            )
            if r.status_code != 200: return []
            items = r.json()
            results = []
            for item in (items if isinstance(items, list) else []):
                s = item.get("smiles", "")
                if not s: continue
                zinc_id = item.get("zinc_id", "")
                results.append({
                    "name": item.get("name") or zinc_id,
                    "smiles": s,
                    "formula": "",
                    "mw": str(item.get("mwt", "")),
                    "xlogp": str(item.get("logp", "")),
                    "tpsa": str(item.get("tpsa", "")),
                    "source": "ZINC",
                    "source_color": "#10b981",
                    "ext_id": zinc_id,
                    "purchasable": True,
                    "pubchem_url": f"https://zinc.docking.org/substances/{zinc_id}/",
                    "structure_url": f"https://zinc.docking.org/substances/{zinc_id}/image.png",
                })
            return results
        except Exception as e:
            print(f"ZINC similarity error: {e}"); return []

    # Run all three in parallel
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        pubchem_res, chembl_res, zinc_res = await asyncio.gather(
            pubchem_similarity(client),
            chembl_similarity(client),
            zinc_similarity(client),
            return_exceptions=True
        )

    # Merge and deduplicate by SMILES
    all_results = []
    seen_smiles = set()
    for source_results in [pubchem_res, chembl_res, zinc_res]:
        if isinstance(source_results, list):
            for r in source_results:
                s = r.get("smiles", "")
                if s and s not in seen_smiles:
                    seen_smiles.add(s)
                    all_results.append(r)

    # Sort: ChEMBL approved drugs first, then by source
    source_order = {"ChEMBL": 0, "PubChem": 1, "ZINC": 2}
    all_results.sort(key=lambda x: source_order.get(x.get("source",""), 3))

    sources_used = list({r.get("source") for r in all_results})
    return {
        "results": all_results[:max_res],
        "count": len(all_results[:max_res]),
        "query_smiles": smiles,
        "sources": sources_used,
        "threshold": threshold / 100,
    }


# ══════════════════════════════════════════════════════════════════════════════
# ADVANCED ANALYSIS ENDPOINTS  (Scaffold · PAINS · Targets · Lead Opt)
# ══════════════════════════════════════════════════════════════════════════════

class AdvancedRequest(BaseModel):
    smiles: str
    batch: Optional[List[str]] = None   # for scaffold batch comparison

# ── PAINS alerts (subset of Baell & Holloway 2010, RDKit-compatible) ──────────
PAINS_ALERTS = [
    ("Quinone", "O=C1C=CC(=O)C=C1"),
    ("Catechol", "Oc1ccccc1O"),
    ("Michael acceptor (vinyl ketone)", "C=CC(=O)"),
    ("Aldehyde", "[CX3H1](=O)"),
    ("Thiol reactive (isothiocyanate)", "N=C=S"),
    ("Epoxide", "C1OC1"),
    ("Acylhydrazide", "C(=O)NN"),
    ("Sulfonyl halide", "S(=O)(=O)[F,Cl,Br,I]"),
    ("Acrylate/acrylamide", "C=CC(=O)[O,N]"),
    ("Rhodanine", "O=C1CSC(=S)N1"),
    ("Hydroxamic acid", "C(=O)NO"),
    ("Azide", "N=[N+]=[N-]"),
    ("Diazo", "N#N"),
    ("Nitroso", "N=O"),
    ("Nitro aromatic", "[nX2]1ccccc1[N+](=O)[O-]"),
    ("Polyhalide", "[F,Cl,Br,I][CX4][F,Cl,Br,I]"),
    ("Imine reactive", "C=N[OH]"),
    ("Thiazolidinedione", "O=C1CSC(=O)N1"),
    ("Mannich base", "N[CH2]C(=O)"),
    ("Coumarin", "O=C1OC2=CC=CC=C2C=C1"),
    ("Propargylamine", "C#CCN"),
    ("Anilinic NH2 (aniline)", "Nc1ccccc1"),
    ("Schiff base", "C=N[!H]"),
    ("Beta-lactam (assay interference)", "C1CC(=O)N1"),
    ("Disulfide", "SSC"),
    ("Peroxide", "OO"),
    ("Heavy metal chelator (EDTA-like)", "OCC(=O)NCC(=O)O"),
    ("Iodoacetamide", "ICC(=O)N"),
    ("Maleimide", "O=C1C=CC(=O)N1"),
    ("Activated ester", "C(=O)ON"),
]

# ── Drug-target pharmacophore patterns ────────────────────────────────────────
TARGET_PATTERNS = [
    {
        "target": "Serine Proteases",
        "family": "Hydrolases",
        "smarts": ["C(=O)[NH]C(=O)", "C1CC(=O)N1", "C(=O)c1ccc(cc1)"],
        "examples": ["Aspirin", "Clopidogrel", "Rivaroxaban"],
        "description": "Molecules with electrophilic carbonyls that covalently modify Ser residues",
        "emoji": "✂️"
    },
    {
        "target": "Kinases (ATP-site)",
        "family": "Transferases",
        "smarts": ["c1cnc2[nH]cnc2c1", "c1ccc2[nH]cnc2c1", "Nc1ncnc2[nH]cnc12", "c1cnc(N)nc1"],
        "examples": ["Imatinib", "Gefitinib", "Erlotinib"],
        "description": "Adenine-like scaffolds that compete with ATP in kinase active sites",
        "emoji": "🔑"
    },
    {
        "target": "GPCRs (Aminergic)",
        "family": "G-Protein Coupled Receptors",
        "smarts": ["NCCc1ccccc1", "NCCC1=CC=CC=C1", "NCC1=CC=CC=C1"],
        "examples": ["Dopamine", "Adrenaline", "Propranolol"],
        "description": "Phenethylamine scaffold typical of aminergic GPCR ligands",
        "emoji": "📡"
    },
    {
        "target": "Nuclear Receptors",
        "family": "Transcription Factors",
        "smarts": ["[C@@H]1CC[C@H]2", "C1CCC2CCCCC2C1", "OC1=CC2=CCCC2CC1"],
        "examples": ["Estradiol", "Testosterone", "Dexamethasone"],
        "description": "Steroidal scaffold binding to nuclear hormone receptors",
        "emoji": "🧬"
    },
    {
        "target": "COX-1/COX-2 (NSAIDs)",
        "family": "Oxidoreductases",
        "smarts": ["c1ccc(cc1)C(=O)O", "CC(=O)Oc1ccccc1", "c1ccc(cc1)S(=O)(=O)N"],
        "examples": ["Aspirin", "Ibuprofen", "Celecoxib"],
        "description": "Carboxylic acid or sulfonamide groups on aryl scaffolds typical of NSAIDs",
        "emoji": "🔥"
    },
    {
        "target": "ACE / Metalloproteinases",
        "family": "Metalloproteases",
        "smarts": ["C(=O)CC(N)C(=O)O", "C(=O)N1CCCC1C(=O)O", "SCC(N)C(=O)O"],
        "examples": ["Lisinopril", "Captopril", "Enalapril"],
        "description": "Zinc-chelating groups (thiol, carboxylate) targeting metalloprotease active sites",
        "emoji": "⚙️"
    },
    {
        "target": "Topoisomerase / DNA intercalators",
        "family": "Isomerases",
        "smarts": ["c1ccc2cc3ccccc3cc2c1", "c1ccc2ccccc2c1", "O=C1c2ccccc2C(=O)c2ccccc21"],
        "examples": ["Doxorubicin", "Daunorubicin", "Mitoxantrone"],
        "description": "Polycyclic aromatic systems that intercalate between DNA base pairs",
        "emoji": "🧪"
    },
    {
        "target": "Cholinesterases (AChE/BChE)",
        "family": "Hydrolases",
        "smarts": ["OC(=O)N(C)C", "OC(=O)c1ccccc1", "[N+](C)(C)(C)CC"],
        "examples": ["Donepezil", "Rivastigmine", "Galantamine"],
        "description": "Carbamate or quaternary ammonium groups targeting cholinesterase active site",
        "emoji": "🧠"
    },
    {
        "target": "HMG-CoA Reductase (Statins)",
        "family": "Oxidoreductases",
        "smarts": ["OCC(O)CC(=O)O", "OC(CC(=O)O)CC(O)", "C1CC(O)CC(=O)O1"],
        "examples": ["Atorvastatin", "Simvastatin", "Rosuvastatin"],
        "description": "Dihydroxy acid pharmacophore mimicking HMG-CoA substrate",
        "emoji": "💊"
    },
    {
        "target": "Beta-Lactamases / PBPs",
        "family": "Antibacterials",
        "smarts": ["C1CC(=O)N1", "C1CN2C(=O)C(C2=O)", "S1CC2N(C1)C(=O)"],
        "examples": ["Penicillin", "Amoxicillin", "Cephalosporin"],
        "description": "Beta-lactam ring that acylates the active-site serine of penicillin-binding proteins",
        "emoji": "🦠"
    },
    {
        "target": "Tubulin / Microtubule",
        "family": "Cytoskeletal",
        "smarts": ["c1ccc(cc1)-c1ccc(cc1)", "C(=O)Nc1ccccc1OC", "c1cncc(c1)C(=O)N"],
        "examples": ["Paclitaxel", "Colchicine", "Vincristine"],
        "description": "Colchicine-site or taxane-site binders disrupting microtubule dynamics",
        "emoji": "🌀"
    },
    {
        "target": "DHFR / Antifolates",
        "family": "Reductases",
        "smarts": ["Nc1nc(N)c2ncc(Cc3ccccc3)nc2n1", "Nc1nc2ccc(CNC3=CC=C(C=C3)C(=O)N)cc2s1"],
        "examples": ["Methotrexate", "Trimethoprim", "Pyrimethamine"],
        "description": "2,4-diaminopyrimidine scaffold targeting dihydrofolate reductase",
        "emoji": "🧫"
    },
    {
        "target": "Ion Channels (Na⁺/Ca²⁺)",
        "family": "Ion Channels",
        "smarts": ["c1ccccc1N(C)C", "CNCC(O)c1ccccc1", "CC(N)Cc1ccccc1"],
        "examples": ["Lidocaine", "Verapamil", "Amlodipine"],
        "description": "Tertiary amine + aromatic pharmacophore blocking ion channel pores",
        "emoji": "⚡"
    },
    {
        "target": "Dihydropteroate Synthase (Sulfonamides)",
        "family": "Antibacterials",
        "smarts": ["c1ccc(cc1)S(=O)(=O)N", "NS(=O)(=O)c1ccccc1"],
        "examples": ["Sulfamethoxazole", "Sulfadiazine"],
        "description": "Sulfonamide group mimicking PABA substrate of bacterial DHPS",
        "emoji": "🔬"
    },
    {
        "target": "PDE Inhibitors",
        "family": "Phosphodiesterases",
        "smarts": ["c1nc2[nH]cnc2c(=O)[nH]1", "O=c1[nH]cnc2ncnc12", "Cc1nc2c([nH]1)c(=O)[nH]c(=O)n2"],
        "examples": ["Sildenafil", "Tadalafil", "Theophylline"],
        "description": "Purine-like scaffold occupying the cyclic nucleotide binding pocket of PDEs",
        "emoji": "💡"
    },
]

@app.post("/scaffold")
async def scaffold_analysis(req: AdvancedRequest):
    """Extract Murcko scaffold, ring systems, and batch scaffold comparison"""
    if not RDKIT_OK:
        return {"error": "RDKit not available"}
    try:
        from rdkit.Chem.Scaffolds import MurckoScaffold
        from rdkit.Chem import Draw, rdDepictor
        from rdkit.Chem import rdMolDescriptors

        mol = Chem.MolFromSmiles(req.smiles.strip())
        if mol is None:
            return {"error": "Invalid SMILES"}

        # ── Core Murcko scaffold ─────────────────────────────────────────────
        scaffold = MurckoScaffold.GetScaffoldForMol(mol)
        scaffold_smiles = Chem.MolToSmiles(scaffold) if scaffold else ""

        # Generic scaffold (all side chains removed, heteroatoms → C)
        generic = MurckoScaffold.MakeScaffoldGeneric(scaffold) if scaffold else None
        generic_smiles = Chem.MolToSmiles(generic) if generic else ""

        # ── Ring systems ─────────────────────────────────────────────────────
        ring_info = mol.GetRingInfo()
        rings = ring_info.AtomRings()
        ring_count = len(rings)
        ring_sizes = sorted(set(len(r) for r in rings))

        # Identify ring types
        ring_types = []
        for ring in rings:
            atoms = [mol.GetAtomWithIdx(i) for i in ring]
            is_aromatic = all(a.GetIsAromatic() for a in atoms)
            has_hetero = any(a.GetAtomicNum() not in (6,1) for a in atoms)
            size = len(ring)
            label = f"{'Aromatic' if is_aromatic else 'Aliphatic'} {'heterocyclic' if has_hetero else 'carbocyclic'} {size}-membered"
            ring_types.append(label)

        # ── Scaffold atoms (which atoms are in scaffold) ──────────────────────
        scaffold_atom_indices = []
        if scaffold:
            match = mol.GetSubstructMatch(scaffold)
            scaffold_atom_indices = list(match)

        # ── Scaffold MW and formula ───────────────────────────────────────────
        scaffold_props = {}
        if scaffold and scaffold.GetNumAtoms() > 0:
            scaffold_props = {
                "mw": round(Descriptors.MolWt(scaffold), 2),
                "formula": rdMolDescriptors.CalcMolFormula(scaffold),
                "heavy_atoms": scaffold.GetNumHeavyAtoms(),
                "rings": scaffold.GetRingInfo().NumRings(),
            }

        # ── Batch scaffold comparison ─────────────────────────────────────────
        batch_results = []
        if req.batch:
            scaffold_groups = {}
            for smi in req.batch:
                smi = smi.strip()
                if not smi:
                    continue
                m = Chem.MolFromSmiles(smi)
                if m is None:
                    batch_results.append({"smiles": smi, "error": "Invalid SMILES"})
                    continue
                try:
                    sc = MurckoScaffold.GetScaffoldForMol(m)
                    sc_smi = Chem.MolToSmiles(sc) if sc else "No scaffold"
                    shared = sc_smi == scaffold_smiles
                    batch_results.append({
                        "smiles": smi,
                        "scaffold": sc_smi,
                        "shared_scaffold": shared,
                        "ring_count": m.GetRingInfo().NumRings(),
                    })
                    scaffold_groups[sc_smi] = scaffold_groups.get(sc_smi, 0) + 1
                except:
                    batch_results.append({"smiles": smi, "error": "Processing failed"})

        return {
            "scaffold_smiles": scaffold_smiles,
            "generic_scaffold_smiles": generic_smiles,
            "scaffold_atom_indices": scaffold_atom_indices,
            "scaffold_props": scaffold_props,
            "ring_count": ring_count,
            "ring_sizes": ring_sizes,
            "ring_types": list(set(ring_types)),
            "framework_atoms": len(scaffold_atom_indices),
            "side_chain_atoms": mol.GetNumHeavyAtoms() - len(scaffold_atom_indices),
            "batch_results": batch_results,
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/pains")
async def pains_filter(req: AdvancedRequest):
    """Screen molecule against PAINS alerts"""
    if not RDKIT_OK:
        return {"error": "RDKit not available"}
    try:
        # Also use RDKit's built-in FilterCatalog for PAINS
        from rdkit.Chem.FilterCatalog import FilterCatalog, FilterCatalogParams

        mol = Chem.MolFromSmiles(req.smiles.strip())
        if mol is None:
            return {"error": "Invalid SMILES"}

        # ── RDKit built-in PAINS catalog (most comprehensive) ─────────────────
        params = FilterCatalogParams()
        params.AddCatalog(FilterCatalogParams.FilterCatalogs.PAINS_A)
        params.AddCatalog(FilterCatalogParams.FilterCatalogs.PAINS_B)
        params.AddCatalog(FilterCatalogParams.FilterCatalogs.PAINS_C)
        catalog = FilterCatalog(params)

        hits = []
        entries = catalog.GetMatches(mol)
        for entry in entries:
            hits.append({
                "name": entry.GetDescription(),
                "source": "PAINS (Baell & Holloway 2010)",
                "severity": "fail",
                "matched_atoms": [],
            })

        # ── Additional custom alerts ──────────────────────────────────────────
        custom_hits = []
        for name, smarts in PAINS_ALERTS:
            try:
                patt = Chem.MolFromSmarts(smarts)
                if patt and mol.HasSubstructMatch(patt):
                    match_atoms = list(mol.GetSubstructMatch(patt))
                    custom_hits.append({
                        "name": name,
                        "smarts": smarts,
                        "source": "Custom alerts",
                        "severity": "warn",
                        "matched_atoms": match_atoms,
                    })
            except:
                continue

        all_hits = hits + custom_hits
        total = len(all_hits)
        pains_count = len(hits)

        if total == 0:
            verdict = "clean"
            verdict_text = "No PAINS alerts — compound looks clean for HTS"
            color = "green"
        elif pains_count > 0:
            verdict = "fail"
            verdict_text = f"{pains_count} PAINS alert(s) detected — likely to cause assay interference"
            color = "red"
        else:
            verdict = "warn"
            verdict_text = f"{len(custom_hits)} structural alert(s) — use with caution in HTS"
            color = "yellow"

        return {
            "verdict": verdict,
            "verdict_text": verdict_text,
            "color": color,
            "pains_hits": hits,
            "custom_hits": custom_hits,
            "total_alerts": total,
            "pains_count": pains_count,
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/targets")
async def target_prediction(req: AdvancedRequest):
    """Predict likely drug targets using pharmacophore SMARTS patterns"""
    if not RDKIT_OK:
        return {"error": "RDKit not available"}
    try:
        mol = Chem.MolFromSmiles(req.smiles.strip())
        if mol is None:
            return {"error": "Invalid SMILES"}

        matches = []
        for tp in TARGET_PATTERNS:
            hit_count = 0
            total = len(tp["smarts"])
            matched_smarts = []
            for smarts in tp["smarts"]:
                try:
                    patt = Chem.MolFromSmarts(smarts)
                    if patt and mol.HasSubstructMatch(patt):
                        hit_count += 1
                        matched_smarts.append(smarts)
                except:
                    continue
            if hit_count > 0:
                confidence = round((hit_count / total) * 100)
                # Boost confidence if multiple patterns match
                if hit_count >= 2:
                    confidence = min(confidence + 20, 95)
                matches.append({
                    "target": tp["target"],
                    "family": tp["family"],
                    "confidence": confidence,
                    "patterns_matched": hit_count,
                    "patterns_total": total,
                    "examples": tp["examples"],
                    "description": tp["description"],
                    "emoji": tp["emoji"],
                    "matched_smarts": matched_smarts,
                })

        # Sort by confidence
        matches.sort(key=lambda x: x["confidence"], reverse=True)

        return {
            "targets": matches,
            "count": len(matches),
            "top_target": matches[0] if matches else None,
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/leadopt")
async def lead_optimisation(req: AdvancedRequest):
    """Suggest lead optimisation changes to improve drug-likeness"""
    if not RDKIT_OK:
        return {"error": "RDKit not available"}
    try:
        from rdkit.Chem import AllChem, rdMolTransforms

        mol = Chem.MolFromSmiles(req.smiles.strip())
        if mol is None:
            return {"error": "Invalid SMILES"}

        # ── Calculate current properties ─────────────────────────────────────
        mw     = Descriptors.MolWt(mol)
        logp   = Crippen.MolLogP(mol)
        hbd    = rdmd.CalcNumHBD(mol)
        hba    = rdmd.CalcNumHBA(mol)
        tpsa   = rdmd.CalcTPSA(mol)
        rotb   = rdmd.CalcNumRotatableBonds(mol)
        qed    = QED.qed(mol)
        rings  = mol.GetRingInfo().NumRings()
        arom   = rdmd.CalcNumAromaticRings(mol)
        heavy  = mol.GetNumHeavyAtoms()

        issues = []
        suggestions = []
        score_deductions = 0

        # ── Rule checks and suggestions ───────────────────────────────────────
        if mw > 500:
            excess = round(mw - 500)
            issues.append({"param": "Molecular Weight", "value": round(mw,1), "target": "≤500 Da", "severity": "high" if mw>600 else "medium"})
            suggestions.append({
                "type": "reduce_mw",
                "priority": "high" if mw>600 else "medium",
                "title": f"Reduce MW by ~{excess} Da",
                "detail": "Remove or truncate bulky substituents. Consider replacing aromatic rings with smaller heteroaromatic systems, or trimming alkyl chains.",
                "impact": f"MW {round(mw,1)} → target <500 Da",
                "icon": "⚖️"
            })
            score_deductions += 2 if mw > 600 else 1

        if logp > 5:
            issues.append({"param": "LogP", "value": round(logp,2), "target": "≤5", "severity": "high" if logp>7 else "medium"})
            suggestions.append({
                "type": "reduce_logp",
                "priority": "high" if logp>7 else "medium",
                "title": f"Reduce lipophilicity (LogP {round(logp,1)} → <5)",
                "detail": "Add polar groups (OH, NH₂, COOH) to reduce LogP. Replace lipophilic aromatic rings with pyridine/pyrimidine. Remove halogen substituents.",
                "impact": "Improves aqueous solubility and reduces CYP metabolism liability",
                "icon": "💧"
            })
            score_deductions += 2 if logp > 7 else 1
        elif logp < 0:
            issues.append({"param": "LogP", "value": round(logp,2), "target": "0–5", "severity": "low"})
            suggestions.append({
                "type": "increase_logp",
                "priority": "low",
                "title": f"Slightly increase lipophilicity (LogP {round(logp,1)} → 0–2)",
                "detail": "Add small alkyl groups or a fluorine to improve membrane permeability.",
                "impact": "Improves cell membrane penetration",
                "icon": "🔆"
            })
            score_deductions += 1

        if hbd > 5:
            issues.append({"param": "H-Bond Donors", "value": hbd, "target": "≤5", "severity": "medium"})
            suggestions.append({
                "type": "reduce_hbd",
                "priority": "medium",
                "title": f"Reduce H-bond donors ({hbd} → ≤5)",
                "detail": "Cap NH or OH groups via methylation or acetylation. Consider prodrug strategies for essential polar groups.",
                "impact": "Improves oral bioavailability and membrane permeability",
                "icon": "🔗"
            })
            score_deductions += 1

        if hba > 10:
            issues.append({"param": "H-Bond Acceptors", "value": hba, "target": "≤10", "severity": "medium"})
            suggestions.append({
                "type": "reduce_hba",
                "priority": "medium",
                "title": f"Reduce H-bond acceptors ({hba} → ≤10)",
                "detail": "Replace polyether chains with simpler linkers. Remove excess carbonyl or ether groups.",
                "impact": "Improves GI absorption",
                "icon": "🎯"
            })
            score_deductions += 1

        if tpsa > 140:
            issues.append({"param": "TPSA", "value": round(tpsa,1), "target": "≤140 Ų", "severity": "high" if tpsa>180 else "medium"})
            suggestions.append({
                "type": "reduce_tpsa",
                "priority": "high" if tpsa>180 else "medium",
                "title": f"Reduce TPSA ({round(tpsa,1)} → <140 Ų)",
                "detail": "Methylate polar NH groups, replace carboxylic acids with bioisosteres (tetrazole, sulfonamide with lower TPSA). Avoid adding more N/O atoms.",
                "impact": "Essential for CNS penetration (<90 Ų) and oral absorption (<140 Ų)",
                "icon": "🌐"
            })
            score_deductions += 2 if tpsa > 180 else 1

        if rotb > 10:
            issues.append({"param": "Rotatable Bonds", "value": rotb, "target": "≤10", "severity": "medium"})
            suggestions.append({
                "type": "reduce_rotb",
                "priority": "medium",
                "title": f"Reduce rotatable bonds ({rotb} → ≤10)",
                "detail": "Cyclise flexible chains into rings. Replace long alkyl linkers with shorter, constrained linkers. Use macrocyclisation for very flexible molecules.",
                "impact": "Improves oral bioavailability and reduces entropic cost of binding",
                "icon": "🔄"
            })
            score_deductions += 1

        if arom > 3:
            issues.append({"param": "Aromatic Rings", "value": arom, "target": "≤3", "severity": "medium"})
            suggestions.append({
                "type": "reduce_arom",
                "priority": "low",
                "title": f"Reduce aromatic ring count ({arom} → ≤3)",
                "detail": "Replace one aromatic ring with a saturated ring (sp3 enrichment). This improves aqueous solubility and reduces flat-molecule aggregation.",
                "impact": "Reduces promiscuity, improves selectivity and solubility",
                "icon": "⭕"
            })
            score_deductions += 1

        # ── Bioisostere suggestions ───────────────────────────────────────────
        bioisosteres = []

        # Carboxylic acid → tetrazole
        cooh = Chem.MolFromSmarts("C(=O)O")
        if mol.HasSubstructMatch(cooh):
            bioisosteres.append({
                "original": "Carboxylic acid (-COOH)",
                "replacement": "Tetrazole (-CN₄H)",
                "reason": "Same pKa, better membrane permeability, metabolically more stable",
                "icon": "🔄"
            })

        # Amide → bioisosteres
        amide = Chem.MolFromSmarts("C(=O)N")
        if mol.HasSubstructMatch(amide):
            bioisosteres.append({
                "original": "Amide (-CONH-)",
                "replacement": "Oxazole or (E)-alkene",
                "reason": "Resists proteolytic cleavage, maintains H-bonding geometry",
                "icon": "🔄"
            })

        # Phenol → fluorophenyl
        phenol = Chem.MolFromSmarts("Oc1ccccc1")
        if mol.HasSubstructMatch(phenol):
            bioisosteres.append({
                "original": "Phenol (-OH on ring)",
                "replacement": "Fluorine (-F) or difluoromethyl (-CHF₂)",
                "reason": "Reduces glucuronidation liability, maintains H-bond character",
                "icon": "🔄"
            })

        # ── Overall score ─────────────────────────────────────────────────────
        base_score = round(qed * 100)
        opt_potential = min(score_deductions * 12, 40)

        priority_order = {"high": 0, "medium": 1, "low": 2}
        suggestions.sort(key=lambda x: priority_order.get(x["priority"], 3))

        return {
            "current_properties": {
                "mw": round(mw, 1), "logp": round(logp, 2),
                "hbd": hbd, "hba": hba, "tpsa": round(tpsa, 1),
                "rotb": rotb, "arom": arom, "qed": round(qed, 3),
            },
            "issues": issues,
            "suggestions": suggestions,
            "bioisosteres": bioisosteres,
            "qed_score": round(qed, 3),
            "optimisation_potential": opt_potential,
            "lipinski_pass": mw<=500 and logp<=5 and hbd<=5 and hba<=10,
            "issue_count": len(issues),
        }
    except Exception as e:
        return {"error": str(e)}
